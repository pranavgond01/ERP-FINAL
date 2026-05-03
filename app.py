from flask import Flask, request, redirect, session, flash, render_template_string, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from datetime import datetime
from io import BytesIO
import sqlite3

app = Flask(__name__)
app.secret_key = "change_this_secret_key"
DB = "plit_erp.db"

COLLEGE = {
    "name": "Pankaj Laddhad Institute of Technology & Management Studies",
    "short": "PLITMS",
    "tagline": "Striving for Excellence in the Quality Professional Education",
    "address": "Chikhli Road, Yelgaon, Buldhana - 443002, Maharashtra",
    "logo": "https://plit.ac.in/wp-content/uploads/2024/01/PLIT-HEader-1-1024x81.jpg"
}


def db():
    con = sqlite3.connect(DB)
    con.row_factory = sqlite3.Row
    return con


def now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def login_required(roles=None):
    def deco(fn):
        @wraps(fn)
        def wrap(*args, **kwargs):
            if "role" not in session:
                return redirect("/")
            if roles and session["role"] not in roles:
                flash("Access denied", "danger")
                return redirect("/")
            return fn(*args, **kwargs)
        return wrap
    return deco


def init_db():
    con = db()
    c = con.cursor()

    c.execute("""CREATE TABLE IF NOT EXISTS users(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        email TEXT UNIQUE,
        password TEXT,
        role TEXT,
        branch TEXT,
        created_at TEXT
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS courses(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE,
        duration INTEGER
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS subjects(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS students(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        name TEXT,
        email TEXT UNIQUE,
        roll_no TEXT,
        course TEXT,
        year TEXT,
        semester TEXT,
        admission_year INTEGER,
        passout_year INTEGER,
        phone TEXT,
        address TEXT,
        fees_total INTEGER,
        fees_paid INTEGER,
        created_at TEXT
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS teachers(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        name TEXT,
        email TEXT UNIQUE,
        course TEXT,
        subject TEXT,
        phone TEXT,
        qualification TEXT,
        lectures_taken INTEGER,
        created_at TEXT
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS teacher_subjects(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        teacher_id INTEGER,
        course TEXT,
        semester TEXT,
        subject TEXT,
        created_at TEXT
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS student_subjects(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER,
        subject TEXT,
        teacher_name TEXT,
        created_at TEXT
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS marks(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER,
        subject TEXT,
        test_name TEXT,
        marks_obtained INTEGER,
        total_marks INTEGER,
        created_at TEXT
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS attendance(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER,
        subject TEXT,
        total_lectures INTEGER,
        present_lectures INTEGER,
        created_at TEXT
    )""")

    courses = [
        ("Computer Science and Engineering", 4),
        ("Civil Engineering", 4),
        ("Mechanical Engineering", 4),
        ("Electrical Engineering", 4),
        ("Electronics and Telecommunication Engineering", 4),
        ("MBA", 2)
    ]

    subjects = [
        "Database Management System",
        "Software Engineering",
        "Design and Analysis of Algorithms",
        "Computer Networks",
        "Operating System",
        "Cyber Security",
        "Web Technology",
        "Engineering Mathematics",
        "Data Structures",
        "Artificial Intelligence",
        "Machine Learning"
    ]

    for course, duration in courses:
        c.execute("INSERT OR IGNORE INTO courses(name,duration) VALUES(?,?)", (course, duration))

    for subject in subjects:
        c.execute("INSERT OR IGNORE INTO subjects(name) VALUES(?)", (subject,))

    default_users = [
        ("Admin", "admin@plit.ac.in", "admin123", "admin", "All"),
        ("Principal", "principal@plit.ac.in", "principal123", "principal", "All"),
        ("CSE HOD", "hod@plit.ac.in", "hod123", "hod", "Computer Science and Engineering"),
        ("Demo Teacher", "teacher@plit.ac.in", "teacher123", "teacher", "Computer Science and Engineering"),
        ("Demo Student", "student@plit.ac.in", "student123", "student", "Computer Science and Engineering")
    ]

    for name, email, password, role, branch in default_users:
        if not c.execute("SELECT * FROM users WHERE email=?", (email,)).fetchone():
            c.execute(
                "INSERT INTO users(name,email,password,role,branch,created_at) VALUES(?,?,?,?,?,?)",
                (name, email, generate_password_hash(password), role, branch, now())
            )

    teacher_user = c.execute("SELECT * FROM users WHERE email='teacher@plit.ac.in'").fetchone()
    if teacher_user and not c.execute("SELECT * FROM teachers WHERE email='teacher@plit.ac.in'").fetchone():
        c.execute("""INSERT INTO teachers(user_id,name,email,course,subject,phone,qualification,lectures_taken,created_at)
        VALUES(?,?,?,?,?,?,?,?,?)""",
        (teacher_user["id"], "Demo Teacher", "teacher@plit.ac.in", "Computer Science and Engineering",
         "Database Management System", "9876543210", "M.Tech", 25, now()))

    student_user = c.execute("SELECT * FROM users WHERE email='student@plit.ac.in'").fetchone()
    if student_user and not c.execute("SELECT * FROM students WHERE email='student@plit.ac.in'").fetchone():
        c.execute("""INSERT INTO students(user_id,name,email,roll_no,course,year,semester,admission_year,passout_year,phone,address,fees_total,fees_paid,created_at)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (student_user["id"], "Demo Student", "student@plit.ac.in", "46", "Computer Science and Engineering",
         "3rd Year", "6th Semester", 2023, 2027, "9999999999", "Buldhana", 50000, 30000, now()))

    con.commit()
    con.close()


def get_common():
    con = db()
    courses = con.execute("SELECT * FROM courses ORDER BY name").fetchall()
    subjects = con.execute("SELECT * FROM subjects ORDER BY name").fetchall()
    con.close()
    years = ["1st Year", "2nd Year", "3rd Year", "4th Year"]
    semesters = ["1st Semester", "2nd Semester", "3rd Semester", "4th Semester",
                 "5th Semester", "6th Semester", "7th Semester", "8th Semester"]
    return courses, subjects, years, semesters


def course_duration(course):
    con = db()
    row = con.execute("SELECT duration FROM courses WHERE name=?", (course,)).fetchone()
    con.close()
    return row["duration"] if row else 4


BASE = """
<!DOCTYPE html>
<html>
<head>
<title>PLITMS ERP</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
body{background:#eef3fb;font-family:Arial}
.navbar{background:linear-gradient(90deg,#002b5b,#0066cc)}
.hero{background:linear-gradient(135deg,#002b5b,#0d6efd);color:white;padding:25px 0}
.logo{max-width:900px;width:100%;background:white;border-radius:12px;padding:8px;margin-bottom:12px}
.card{border:0;border-radius:16px;box-shadow:0 8px 22px rgba(0,0,0,.08);margin-bottom:22px}
.stat{background:white;border-left:6px solid #0d6efd;padding:18px;border-radius:14px;box-shadow:0 6px 18px rgba(0,0,0,.08)}
.table th{background:#002b5b;color:white;white-space:nowrap}
.table td{white-space:nowrap}
input,select,textarea,.btn{border-radius:10px!important}
</style>
</head>
<body>
<nav class="navbar navbar-dark">
<div class="container-fluid">
<a class="navbar-brand fw-bold" href="/">PLITMS ERP</a>
{% if session.get('role') %}
<span class="text-white">{{session.get('name')}} | {{session.get('role')|title}} | {{session.get('branch')}}
<a class="btn btn-light btn-sm ms-2" href="/logout">Logout</a></span>
{% endif %}
</div>
</nav>
<div class="hero">
<div class="container">
<img src="{{college.logo}}" class="logo">
<h3>{{college.name}}</h3>
<p>{{college.tagline}}</p>
<small>{{college.address}}</small>
</div>
</div>
<div class="container my-4">
{% with messages=get_flashed_messages(with_categories=true) %}
{% for c,m in messages %}
<div class="alert alert-{{c}}">{{m}}</div>
{% endfor %}
{% endwith %}
{{content|safe}}
</div>
</body>
</html>
"""


def page(content):
    return render_template_string(BASE, content=content, college=COLLEGE)


@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form["email"].strip()
        password = request.form["password"].strip()

        con = db()
        user = con.execute("SELECT * FROM users WHERE email=?", (email,)).fetchone()
        con.close()

        if user and check_password_hash(user["password"], password):
            session["user_id"] = user["id"]
            session["name"] = user["name"]
            session["email"] = user["email"]
            session["role"] = user["role"]
            session["branch"] = user["branch"]
            return redirect("/" + user["role"])

        flash("Invalid login", "danger")

    return page("""
    <div class="row justify-content-center">
    <div class="col-md-5">
    <div class="card p-4">
    <h3 class="text-center text-primary">Login</h3>
    <p class="text-center text-muted">Student | Teacher | HOD | Principal | Admin</p>
    <form method="POST">
    <input class="form-control mb-3" name="email" placeholder="Email" required>
    <input class="form-control mb-3" name="password" type="password" placeholder="Password" required>
    <button class="btn btn-primary w-100">Login</button>
    </form>
    </div></div></div>
    """)


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")


def filtered_students():
    search = request.args.get("search", "")
    course = request.args.get("course", "")
    year = request.args.get("year", "")
    semester = request.args.get("semester", "")

    if session.get("role") == "hod":
        course = session.get("branch")

    query = "SELECT * FROM students WHERE 1=1"
    params = []

    if search:
        query += " AND (name LIKE ? OR roll_no LIKE ? OR email LIKE ?)"
        key = f"%{search}%"
        params.extend([key, key, key])

    if course:
        query += " AND course=?"
        params.append(course)

    if year:
        query += " AND year=?"
        params.append(year)

    if semester:
        query += " AND semester=?"
        params.append(semester)

    con = db()
    rows = con.execute(query + " ORDER BY CAST(roll_no AS INTEGER) ASC", params).fetchall()
    con.close()
    return rows


def filtered_teachers():
    course = request.args.get("course", "")
    search = request.args.get("search", "")

    if session.get("role") == "hod":
        course = session.get("branch")

    query = "SELECT * FROM teachers WHERE 1=1"
    params = []

    if course:
        query += " AND course=?"
        params.append(course)

    if search:
        query += " AND (name LIKE ? OR email LIKE ? OR subject LIKE ?)"
        key = f"%{search}%"
        params.extend([key, key, key])

    con = db()
    rows = con.execute(query + " ORDER BY id DESC", params).fetchall()
    con.close()
    return rows


def filter_box(role):
    courses, subjects, years, semesters = get_common()
    course = request.args.get("course", "")
    year = request.args.get("year", "")
    semester = request.args.get("semester", "")
    search = request.args.get("search", "")
    qs = request.query_string.decode()

    if role == "hod":
        course_select = f"<input class='form-control' value='{session.get('branch')}' disabled>"
    else:
        co = "".join([f"<option {'selected' if course==c['name'] else ''}>{c['name']}</option>" for c in courses])
        course_select = f"<select name='course' class='form-control'><option value=''>All Branch</option>{co}</select>"

    yo = "".join([f"<option {'selected' if year==y else ''}>{y}</option>" for y in years])
    so = "".join([f"<option {'selected' if semester==s else ''}>{s}</option>" for s in semesters])

    return f"""
    <div class="card p-3">
    <h5>Search / Filter</h5>
    <form method="GET" class="row g-2">
    <div class="col-md-3"><input name="search" value="{search}" class="form-control" placeholder="Search name / roll / email"></div>
    <div class="col-md-3">{course_select}</div>
    <div class="col-md-2"><select name="year" class="form-control"><option value="">All Year</option>{yo}</select></div>
    <div class="col-md-2"><select name="semester" class="form-control"><option value="">All Semester</option>{so}</select></div>
    <div class="col-md-2"><button class="btn btn-primary w-100">Search</button></div>
    <div class="col-md-12 mt-2">
    <a href='/{role}' class='btn btn-secondary'>Reset</a>
    <a href='/export_students_excel?{qs}' class='btn btn-success'>Export Excel</a>
    <a href='/export_students_pdf?{qs}' class='btn btn-danger'>Export PDF</a>
    </div>
    </form>
    </div>
    """


def graph_box(a, b, canvas_id):
    return f"""
    <div class="card p-3">
    <h5>Graph Dashboard</h5>
    <canvas id="{canvas_id}" height="90"></canvas>
    </div>
    <script>
    new Chart(document.getElementById('{canvas_id}'), {{
        type: 'bar',
        data: {{
            labels: ['Students', 'Teachers / Subjects'],
            datasets: [{{ label:'Count', data:[{a},{b}], borderWidth:1 }}]
        }}
    }});
    </script>
    """


def import_box():
    return """
    <div class="card p-3">
    <h5>Import Students From Excel</h5>
    <p>Columns: Name, Email, Password, Roll No, Course, Year, Semester, Admission Year, Phone, Address, Fees Total, Fees Paid</p>
    <form action="/import_students_excel" method="POST" enctype="multipart/form-data">
    <input type="file" name="file" class="form-control mb-2" accept=".xlsx" required>
    <button class="btn btn-warning">Import Excel</button>
    </form></div>
    """


def hod_form():
    courses, subjects, years, semesters = get_common()
    course_options = "".join([f"<option>{c['name']}</option>" for c in courses])

    return f"""
    <div class="card p-3">
    <h5>Add Branch-wise HOD</h5>
    <form action="/principal_add_hod" method="POST" class="row g-2">
    <div class="col-md-3"><input name="name" class="form-control" placeholder="HOD Name" required></div>
    <div class="col-md-3"><input name="email" type="email" class="form-control" placeholder="HOD Email" required></div>
    <div class="col-md-3"><input name="password" class="form-control" placeholder="Password" required></div>
    <div class="col-md-3"><select name="branch" class="form-control" required>{course_options}</select></div>
    <div><button class="btn btn-primary">Add Branch HOD</button></div>
    </form></div>
    """


@app.route("/admin")
@login_required(["admin"])
def admin():
    con = db()
    principals = con.execute("SELECT * FROM users WHERE role='principal' ORDER BY id DESC").fetchall()
    hods = con.execute("SELECT * FROM users WHERE role='hod' ORDER BY id DESC").fetchall()
    teachers = con.execute("SELECT * FROM teachers").fetchall()
    students = con.execute("SELECT * FROM students").fetchall()
    courses = con.execute("SELECT * FROM courses ORDER BY name").fetchall()
    subjects = con.execute("SELECT * FROM subjects ORDER BY name").fetchall()
    con.close()

    content = f"""
    <h3>Admin Dashboard</h3>
    <div class="row mb-4">
    <div class="col"><div class="stat">Principals<br><b>{len(principals)}</b></div></div>
    <div class="col"><div class="stat">HODs<br><b>{len(hods)}</b></div></div>
    <div class="col"><div class="stat">Students<br><b>{len(students)}</b></div></div>
    <div class="col"><div class="stat">Teachers<br><b>{len(teachers)}</b></div></div>
    </div>

    <div class="card p-3"><h5>Add Principal / HOD</h5>
    <form action="/admin_add_user" method="POST" class="row g-2">
    <div class="col-md-3"><input name="name" class="form-control" placeholder="Name" required></div>
    <div class="col-md-3"><input name="email" type="email" class="form-control" placeholder="Email" required></div>
    <div class="col-md-3"><input name="password" class="form-control" placeholder="Password" required></div>
    <div class="col-md-2"><select name="role" class="form-control"><option value="principal">Principal</option><option value="hod">HOD</option></select></div>
    <div class="col-md-1"><button class="btn btn-primary w-100">Add</button></div>
    </form></div>

    <div class="card p-3"><h5>Add / Edit Course</h5>
    <form action="/admin_add_course" method="POST" class="row g-2">
    <div class="col-md-6"><input name="name" class="form-control" placeholder="Course Name" required></div>
    <div class="col-md-3"><input name="duration" type="number" class="form-control" placeholder="Duration Years" required></div>
    <div class="col-md-3"><button class="btn btn-success w-100">Save Course</button></div>
    </form></div>

    <div class="card p-3"><h5>Add Subject</h5>
    <form action="/admin_add_subject" method="POST" class="row g-2">
    <div class="col-md-9"><input name="name" class="form-control" placeholder="Subject Name" required></div>
    <div class="col-md-3"><button class="btn btn-success w-100">Add Subject</button></div>
    </form></div>
    """

    content += "<div class='card p-3'><h5>HODs</h5><table class='table table-bordered'><tr><th>Name</th><th>Email</th><th>Branch</th></tr>"
    for h in hods:
        content += f"<tr><td>{h['name']}</td><td>{h['email']}</td><td>{h['branch']}</td></tr>"
    content += "</table></div>"

    content += "<div class='card p-3'><h5>Courses</h5><table class='table table-bordered'><tr><th>Course</th><th>Duration</th></tr>"
    for c in courses:
        content += f"<tr><td>{c['name']}</td><td>{c['duration']} Years</td></tr>"
    content += "</table></div>"

    content += "<div class='card p-3'><h5>Subjects</h5><table class='table table-bordered'><tr><th>Subject</th><th>Delete</th></tr>"
    for s in subjects:
        content += f"<tr><td>{s['name']}</td><td><a class='btn btn-danger btn-sm' href='/delete_subject/{s['id']}'>Delete</a></td></tr>"
    content += "</table></div>"

    return page(content)


@app.route("/principal")
@login_required(["principal"])
def principal():
    return management_dashboard("principal")


@app.route("/hod")
@login_required(["hod"])
def hod():
    return management_dashboard("hod")


def management_dashboard(role):
    students = filtered_students()
    teachers = filtered_teachers()
    courses, subjects, years, semesters = get_common()

    con = db()
    teacher_subjects = con.execute("""
        SELECT teacher_subjects.*, teachers.name AS teacher_name, teachers.email AS teacher_email
        FROM teacher_subjects JOIN teachers ON teacher_subjects.teacher_id=teachers.id
        ORDER BY teacher_subjects.id DESC
    """).fetchall()
    all_teachers = filtered_teachers()
    con.close()

    teacher_options = "".join([f"<option value='{t['id']}'>{t['name']} - {t['email']}</option>" for t in all_teachers])
    course_options = "".join([f"<option>{c['name']}</option>" for c in courses])
    semester_options = "".join([f"<option>{s}</option>" for s in semesters])
    subject_options = "".join([f"<option>{s['name']}</option>" for s in subjects])

    title = "Principal Dashboard" if role == "principal" else f"{session.get('branch')} HOD Dashboard"

    content = f"""
    <h3>{title}</h3>
    <div class="row mb-4">
    <div class="col"><div class="stat">Filtered Students<br><b>{len(students)}</b></div></div>
    <div class="col"><div class="stat">Filtered Teachers<br><b>{len(teachers)}</b></div></div>
    </div>
    {filter_box(role)}
    {graph_box(len(students), len(teachers), role + "Chart")}
    {import_box()}
    {hod_form() if role == "principal" else ""}
    {student_form(courses, years, semesters)}
    {multi_student_box()}
    {teacher_form(courses, subjects)}
    {multi_teacher_box()}

    <div class="card p-3"><h5>Assign Multiple Subjects to Teacher</h5>
    <form action="/assign_teacher_subject" method="POST" class="row g-2">
    <div class="col-md-3"><select name="teacher_id" class="form-control">{teacher_options}</select></div>
    <div class="col-md-3"><select name="course" class="form-control">{course_options}</select></div>
    <div class="col-md-3"><select name="semester" class="form-control">{semester_options}</select></div>
    <div class="col-md-3"><select name="subject" class="form-control">{subject_options}</select></div>
    <div class="col-md-12"><button class="btn btn-success">Assign Subject</button></div>
    </form></div>
    """

    content += students_table(students, True, True)
    content += teachers_table(teachers, True, True)

    content += "<div class='card p-3'><h5>Teacher Subject Assignments</h5><table class='table table-bordered'><tr><th>Teacher</th><th>Email</th><th>Course</th><th>Semester</th><th>Subject</th><th>Action</th></tr>"
    for ts in teacher_subjects:
        if role == "hod" and ts["course"] != session.get("branch"):
            continue
        content += f"<tr><td>{ts['teacher_name']}</td><td>{ts['teacher_email']}</td><td>{ts['course']}</td><td>{ts['semester']}</td><td>{ts['subject']}</td><td><a class='btn btn-danger btn-sm' href='/delete_teacher_subject/{ts['id']}'>Delete</a></td></tr>"
    content += "</table></div>"

    return page(content)


@app.route("/teacher")
@login_required(["teacher"])
def teacher():
    students = filtered_students()
    courses, subjects, years, semesters = get_common()

    con = db()
    teacher_data = con.execute("SELECT * FROM teachers WHERE email=?", (session["email"],)).fetchone()
    assigned = []
    if teacher_data:
        assigned = con.execute("SELECT * FROM teacher_subjects WHERE teacher_id=? ORDER BY semester", (teacher_data["id"],)).fetchall()
    con.close()

    content = f"""
    <h3>Teacher Dashboard</h3>
    <div class="row mb-4">
    <div class="col"><div class="stat">Filtered Students<br><b>{len(students)}</b></div></div>
    <div class="col"><div class="stat">Assigned Subjects<br><b>{len(assigned)}</b></div></div>
    </div>
    {filter_box("teacher")}
    {graph_box(len(students), len(assigned), "teacherChart")}
    {import_box()}
    <div class="card p-3"><h5>My Assigned Subjects</h5>
    <table class="table table-bordered"><tr><th>Course</th><th>Semester</th><th>Subject</th></tr>
    """
    for a in assigned:
        content += f"<tr><td>{a['course']}</td><td>{a['semester']}</td><td>{a['subject']}</td></tr>"
    content += "</table></div>"

    content += student_form(courses, years, semesters)
    content += multi_student_box()
    content += subject_marks_attendance_forms(students, subjects)
    content += students_table(students, True, True)

    return page(content)


@app.route("/student")
@login_required(["student"])
def student():
    con = db()
    s = con.execute("SELECT * FROM students WHERE email=?", (session["email"],)).fetchone()
    subs = con.execute("SELECT * FROM student_subjects WHERE student_id=?", (s["id"],)).fetchall()
    marks = con.execute("SELECT * FROM marks WHERE student_id=?", (s["id"],)).fetchall()
    att = con.execute("SELECT * FROM attendance WHERE student_id=?", (s["id"],)).fetchall()
    con.close()

    content = f"""
    <h3>Student Dashboard</h3>
    <div class="row mb-4">
    <div class="col"><div class="stat">Total Fees<br><b>₹{s['fees_total']}</b></div></div>
    <div class="col"><div class="stat">Paid Fees<br><b>₹{s['fees_paid']}</b></div></div>
    <div class="col"><div class="stat">Remaining<br><b>₹{s['fees_total'] - s['fees_paid']}</b></div></div>
    </div>
    <div class="card p-3">
    <h5>Student Information</h5>
    <p><b>Name:</b> {s['name']}</p>
    <p><b>Email:</b> {s['email']}</p>
    <p><b>Roll:</b> {s['roll_no']}</p>
    <p><b>Course:</b> {s['course']}</p>
    <p><b>Year:</b> {s['year']}</p>
    <p><b>Semester:</b> {s['semester']}</p>
    <p><b>Passout Year:</b> {s['passout_year']}</p>
    </div>
    """

    content += "<div class='card p-3'><h5>Subjects</h5><table class='table table-bordered'><tr><th>Subject</th><th>Teacher</th></tr>"
    for x in subs:
        content += f"<tr><td>{x['subject']}</td><td>{x['teacher_name']}</td></tr>"
    content += "</table></div>"

    content += "<div class='card p-3'><h5>Marks</h5><table class='table table-bordered'><tr><th>Subject</th><th>Test</th><th>Marks</th><th>Total</th></tr>"
    for m in marks:
        content += f"<tr><td>{m['subject']}</td><td>{m['test_name']}</td><td>{m['marks_obtained']}</td><td>{m['total_marks']}</td></tr>"
    content += "</table></div>"

    content += "<div class='card p-3'><h5>Attendance</h5><table class='table table-bordered'><tr><th>Subject</th><th>Total</th><th>Present</th><th>%</th></tr>"
    for a in att:
        per = round((a["present_lectures"] / a["total_lectures"]) * 100, 2) if a["total_lectures"] else 0
        content += f"<tr><td>{a['subject']}</td><td>{a['total_lectures']}</td><td>{a['present_lectures']}</td><td>{per}%</td></tr>"
    content += "</table></div>"
    return page(content)


def student_form(courses, years, semesters):
    co = "".join([f"<option>{c['name']}</option>" for c in courses])
    yo = "".join([f"<option>{y}</option>" for y in years])
    so = "".join([f"<option>{s}</option>" for s in semesters])
    return f"""
    <div class="card p-3"><h5>Add Student</h5>
    <form action="/add_student" method="POST" class="row g-2">
    <div class="col-md-4"><input name="name" class="form-control" placeholder="Name" required></div>
    <div class="col-md-4"><input name="email" type="email" class="form-control" placeholder="Email" required></div>
    <div class="col-md-4"><input name="password" class="form-control" placeholder="Password" required></div>
    <div class="col-md-3"><input name="roll_no" class="form-control" placeholder="Roll No" required></div>
    <div class="col-md-3"><select name="course" class="form-control">{co}</select></div>
    <div class="col-md-3"><select name="year" class="form-control">{yo}</select></div>
    <div class="col-md-3"><select name="semester" class="form-control">{so}</select></div>
    <div class="col-md-3"><input name="admission_year" type="number" class="form-control" placeholder="Admission Year" required></div>
    <div class="col-md-3"><input name="phone" class="form-control" placeholder="Phone" required></div>
    <div class="col-md-3"><input name="fees_total" type="number" class="form-control" placeholder="Total Fees" required></div>
    <div class="col-md-3"><input name="fees_paid" type="number" class="form-control" placeholder="Paid Fees" required></div>
    <div class="col-md-12"><input name="address" class="form-control" placeholder="Address" required></div>
    <div><button class="btn btn-success">Add Student</button></div>
    </form></div>
    """


def multi_student_box():
    return """
    <div class="card p-3"><h5>Add Multiple Students</h5>
    <p>Format: name,email,password,roll,course,year,semester,admission_year,phone,address,fees_total,fees_paid</p>
    <form action="/multi_add_student" method="POST">
    <textarea name="data" rows="4" class="form-control mb-2"></textarea>
    <button class="btn btn-success">Add Multiple Students</button>
    </form></div>
    """


def teacher_form(courses, subjects):
    co = "".join([f"<option>{c['name']}</option>" for c in courses])
    so = "".join([f"<option>{s['name']}</option>" for s in subjects])
    return f"""
    <div class="card p-3"><h5>Add Teacher</h5>
    <form action="/add_teacher" method="POST" class="row g-2">
    <div class="col-md-4"><input name="name" class="form-control" placeholder="Name" required></div>
    <div class="col-md-4"><input name="email" type="email" class="form-control" placeholder="Email" required></div>
    <div class="col-md-4"><input name="password" class="form-control" placeholder="Password" required></div>
    <div class="col-md-4"><select name="course" class="form-control">{co}</select></div>
    <div class="col-md-4"><select name="subject" class="form-control">{so}</select></div>
    <div class="col-md-4"><input name="phone" class="form-control" placeholder="Phone" required></div>
    <div class="col-md-6"><input name="qualification" class="form-control" placeholder="Qualification" required></div>
    <div class="col-md-6"><input name="lectures_taken" type="number" class="form-control" placeholder="Lectures Taken" required></div>
    <div><button class="btn btn-primary">Add Teacher</button></div>
    </form></div>
    """


def multi_teacher_box():
    return """
    <div class="card p-3"><h5>Add Multiple Teachers</h5>
    <p>Format: name,email,password,course,subject,phone,qualification,lectures</p>
    <form action="/multi_add_teacher" method="POST">
    <textarea name="data" rows="4" class="form-control mb-2"></textarea>
    <button class="btn btn-primary">Add Multiple Teachers</button>
    </form></div>
    """


def subject_marks_attendance_forms(students, subjects):
    stu = "".join([f"<option value='{s['id']}'>{s['name']} - {s['roll_no']}</option>" for s in students])
    sub = "".join([f"<option>{x['name']}</option>" for x in subjects])
    return f"""
    <div class="card p-3"><h5>Add Subject / Marks / Attendance</h5>
    <form action="/assign_subject" method="POST" class="row g-2 mb-3">
    <div class="col-md-5"><select name="student_id" class="form-control">{stu}</select></div>
    <div class="col-md-5"><select name="subject" class="form-control">{sub}</select></div>
    <div class="col-md-2"><button class="btn btn-success w-100">Add Subject</button></div>
    </form>
    <form action="/add_marks" method="POST" class="row g-2 mb-3">
    <div class="col-md-3"><select name="student_id" class="form-control">{stu}</select></div>
    <div class="col-md-3"><select name="subject" class="form-control">{sub}</select></div>
    <div class="col-md-2"><input name="test_name" class="form-control" placeholder="Test"></div>
    <div class="col-md-2"><input name="marks_obtained" type="number" class="form-control" placeholder="Marks"></div>
    <div class="col-md-2"><input name="total_marks" type="number" class="form-control" placeholder="Total"></div>
    <div><button class="btn btn-primary">Add Marks</button></div>
    </form>
    <form action="/add_attendance" method="POST" class="row g-2">
    <div class="col-md-3"><select name="student_id" class="form-control">{stu}</select></div>
    <div class="col-md-3"><select name="subject" class="form-control">{sub}</select></div>
    <div class="col-md-3"><input name="total_lectures" type="number" class="form-control" placeholder="Total Lectures"></div>
    <div class="col-md-3"><input name="present_lectures" type="number" class="form-control" placeholder="Present"></div>
    <div><button class="btn btn-warning">Add Attendance</button></div>
    </form></div>
    """


def students_table(students, can_delete=False, can_edit=False):
    html = "<div class='card p-3'><h5>Students</h5><div class='table-responsive'><table class='table table-bordered table-striped'><tr><th>Name</th><th>Email</th><th>Roll</th><th>Course</th><th>Year</th><th>Sem</th><th>Passout</th><th>Remaining</th><th>Action</th></tr>"
    for s in students:
        action = ""
        if can_edit:
            action += f"<a class='btn btn-warning btn-sm me-1' href='/edit_student/{s['id']}'>Edit</a>"
        if can_delete:
            action += f"<a class='btn btn-danger btn-sm' href='/delete_student/{s['id']}'>Delete</a>"
        html += f"<tr><td>{s['name']}</td><td>{s['email']}</td><td>{s['roll_no']}</td><td>{s['course']}</td><td>{s['year']}</td><td>{s['semester']}</td><td>{s['passout_year']}</td><td>₹{s['fees_total']-s['fees_paid']}</td><td>{action}</td></tr>"
    html += "</table></div></div>"
    return html


def teachers_table(teachers, can_delete=False, can_edit=False):
    html = "<div class='card p-3'><h5>Teachers</h5><div class='table-responsive'><table class='table table-bordered table-striped'><tr><th>Name</th><th>Email</th><th>Course</th><th>Subject</th><th>Phone</th><th>Lectures</th><th>Action</th></tr>"
    for t in teachers:
        action = ""
        if can_edit:
            action += f"<a class='btn btn-warning btn-sm me-1' href='/edit_teacher/{t['id']}'>Edit</a>"
        if can_delete:
            action += f"<a class='btn btn-danger btn-sm' href='/delete_teacher/{t['id']}'>Delete</a>"
        html += f"<tr><td>{t['name']}</td><td>{t['email']}</td><td>{t['course']}</td><td>{t['subject']}</td><td>{t['phone']}</td><td>{t['lectures_taken']}</td><td>{action}</td></tr>"
    html += "</table></div></div>"
    return html


@app.route("/add_student", methods=["POST"])
@login_required(["teacher", "principal", "hod"])
def add_student():
    d = request.form
    course = session.get("branch") if session.get("role") == "hod" else d["course"]
    passout = int(d["admission_year"]) + course_duration(course)
    con = db()

    if con.execute("SELECT * FROM users WHERE email=?", (d["email"],)).fetchone():
        flash("Student email already exists", "danger")
        con.close()
        return redirect("/" + session["role"])

    c = con.cursor()
    c.execute("INSERT INTO users(name,email,password,role,branch,created_at) VALUES(?,?,?,?,?,?)",
              (d["name"], d["email"], generate_password_hash(d["password"]), "student", course, now()))
    uid = c.lastrowid

    c.execute("""INSERT INTO students(user_id,name,email,roll_no,course,year,semester,admission_year,passout_year,phone,address,fees_total,fees_paid,created_at)
    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
    (uid, d["name"], d["email"], d["roll_no"], course, d["year"], d["semester"],
     int(d["admission_year"]), passout, d["phone"], d["address"], int(d["fees_total"]), int(d["fees_paid"]), now()))
    con.commit()
    con.close()
    flash("Student added", "success")
    return redirect("/" + session["role"])


@app.route("/multi_add_student", methods=["POST"])
@login_required(["teacher", "principal", "hod"])
def multi_add_student():
    lines = request.form["data"].strip().splitlines()
    count = 0

    for line in lines:
        try:
            name,email,password,roll,course,year,sem,admission,phone,address,total,paid = [x.strip() for x in line.split(",")]
            if session.get("role") == "hod":
                course = session.get("branch")

            passout = int(admission) + course_duration(course)
            con = db()

            if con.execute("SELECT * FROM users WHERE email=?", (email,)).fetchone():
                con.close()
                continue

            c = con.cursor()
            c.execute("INSERT INTO users(name,email,password,role,branch,created_at) VALUES(?,?,?,?,?,?)",
                      (name, email, generate_password_hash(password), "student", course, now()))
            uid = c.lastrowid

            c.execute("""INSERT INTO students(user_id,name,email,roll_no,course,year,semester,admission_year,passout_year,phone,address,fees_total,fees_paid,created_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (uid, name, email, roll, course, year, sem, int(admission), passout, phone, address, int(total), int(paid), now()))
            con.commit()
            con.close()
            count += 1
        except Exception:
            continue

    flash(f"{count} students added", "success")
    return redirect("/" + session["role"])


@app.route("/add_teacher", methods=["POST"])
@login_required(["principal", "admin", "hod"])
def add_teacher():
    d = request.form
    course = session.get("branch") if session.get("role") == "hod" else d["course"]
    con = db()

    if con.execute("SELECT * FROM users WHERE email=?", (d["email"],)).fetchone():
        flash("Teacher email already exists", "danger")
        con.close()
        return redirect("/" + session["role"])

    c = con.cursor()
    c.execute("INSERT INTO users(name,email,password,role,branch,created_at) VALUES(?,?,?,?,?,?)",
              (d["name"], d["email"], generate_password_hash(d["password"]), "teacher", course, now()))
    uid = c.lastrowid

    c.execute("""INSERT INTO teachers(user_id,name,email,course,subject,phone,qualification,lectures_taken,created_at)
    VALUES(?,?,?,?,?,?,?,?,?)""",
    (uid, d["name"], d["email"], course, d["subject"], d["phone"], d["qualification"], int(d["lectures_taken"]), now()))
    con.commit()
    con.close()
    flash("Teacher added", "success")
    return redirect("/" + session["role"])


@app.route("/multi_add_teacher", methods=["POST"])
@login_required(["principal", "admin", "hod"])
def multi_add_teacher():
    lines = request.form["data"].strip().splitlines()
    count = 0

    for line in lines:
        try:
            name,email,password,course,subject,phone,qualification,lectures = [x.strip() for x in line.split(",")]
            if session.get("role") == "hod":
                course = session.get("branch")

            con = db()

            if con.execute("SELECT * FROM users WHERE email=?", (email,)).fetchone():
                con.close()
                continue

            c = con.cursor()
            c.execute("INSERT INTO users(name,email,password,role,branch,created_at) VALUES(?,?,?,?,?,?)",
                      (name, email, generate_password_hash(password), "teacher", course, now()))
            uid = c.lastrowid

            c.execute("""INSERT INTO teachers(user_id,name,email,course,subject,phone,qualification,lectures_taken,created_at)
            VALUES(?,?,?,?,?,?,?,?,?)""",
            (uid, name, email, course, subject, phone, qualification, int(lectures), now()))
            con.commit()
            con.close()
            count += 1
        except Exception:
            continue

    flash(f"{count} teachers added", "success")
    return redirect("/" + session["role"])


@app.route("/principal_add_hod", methods=["POST"])
@login_required(["principal"])
def principal_add_hod():
    d = request.form
    con = db()

    if con.execute("SELECT * FROM users WHERE email=?", (d["email"],)).fetchone():
        flash("HOD email already exists", "danger")
    else:
        con.execute("INSERT INTO users(name,email,password,role,branch,created_at) VALUES(?,?,?,?,?,?)",
                    (d["name"], d["email"], generate_password_hash(d["password"]), "hod", d["branch"], now()))
        con.commit()
        flash("Branch HOD added successfully", "success")

    con.close()
    return redirect("/principal")


@app.route("/assign_teacher_subject", methods=["POST"])
@login_required(["principal", "hod"])
def assign_teacher_subject():
    d = request.form
    course = session.get("branch") if session.get("role") == "hod" else d["course"]

    con = db()
    con.execute("INSERT INTO teacher_subjects(teacher_id,course,semester,subject,created_at) VALUES(?,?,?,?,?)",
                (d["teacher_id"], course, d["semester"], d["subject"], now()))
    con.commit()
    con.close()
    flash("Teacher subject assigned", "success")
    return redirect("/" + session["role"])


@app.route("/delete_teacher_subject/<int:id>")
@login_required(["principal", "hod"])
def delete_teacher_subject(id):
    con = db()
    con.execute("DELETE FROM teacher_subjects WHERE id=?", (id,))
    con.commit()
    con.close()
    flash("Teacher subject deleted", "success")
    return redirect("/" + session["role"])


@app.route("/admin_add_user", methods=["POST"])
@login_required(["admin"])
def admin_add_user():
    d = request.form
    branch = "All"
    if d["role"] == "hod":
        branch = "Computer Science and Engineering"

    con = db()

    if con.execute("SELECT * FROM users WHERE email=?", (d["email"],)).fetchone():
        flash("Email already exists", "danger")
    else:
        con.execute("INSERT INTO users(name,email,password,role,branch,created_at) VALUES(?,?,?,?,?,?)",
                    (d["name"], d["email"], generate_password_hash(d["password"]), d["role"], branch, now()))
        con.commit()
        flash("User added", "success")

    con.close()
    return redirect("/admin")


@app.route("/admin_add_course", methods=["POST"])
@login_required(["admin"])
def admin_add_course():
    d = request.form
    con = db()
    con.execute("""INSERT OR REPLACE INTO courses(id,name,duration)
    VALUES((SELECT id FROM courses WHERE name=?),?,?)""",
    (d["name"], d["name"], int(d["duration"])))
    con.commit()
    con.close()
    flash("Course saved", "success")
    return redirect("/admin")


@app.route("/admin_add_subject", methods=["POST"])
@login_required(["admin"])
def admin_add_subject():
    con = db()
    con.execute("INSERT OR IGNORE INTO subjects(name) VALUES(?)", (request.form["name"],))
    con.commit()
    con.close()
    flash("Subject added", "success")
    return redirect("/admin")


@app.route("/delete_subject/<int:id>")
@login_required(["admin"])
def delete_subject(id):
    con = db()
    con.execute("DELETE FROM subjects WHERE id=?", (id,))
    con.commit()
    con.close()
    flash("Subject deleted", "success")
    return redirect("/admin")


@app.route("/delete_student/<int:id>")
@login_required(["teacher", "principal", "hod"])
def delete_student(id):
    con = db()
    s = con.execute("SELECT * FROM students WHERE id=?", (id,)).fetchone()

    if s:
        if session.get("role") == "hod" and s["course"] != session.get("branch"):
            con.close()
            flash("You can delete only your branch students", "danger")
            return redirect("/hod")

        con.execute("DELETE FROM users WHERE email=?", (s["email"],))
        con.execute("DELETE FROM students WHERE id=?", (id,))
        con.execute("DELETE FROM student_subjects WHERE student_id=?", (id,))
        con.execute("DELETE FROM marks WHERE student_id=?", (id,))
        con.execute("DELETE FROM attendance WHERE student_id=?", (id,))

    con.commit()
    con.close()
    flash("Student deleted", "success")
    return redirect("/" + session["role"])


@app.route("/delete_teacher/<int:id>")
@login_required(["principal", "hod"])
def delete_teacher(id):
    con = db()
    t = con.execute("SELECT * FROM teachers WHERE id=?", (id,)).fetchone()

    if t:
        if session.get("role") == "hod" and t["course"] != session.get("branch"):
            con.close()
            flash("You can delete only your branch teachers", "danger")
            return redirect("/hod")

        con.execute("DELETE FROM users WHERE email=?", (t["email"],))
        con.execute("DELETE FROM teacher_subjects WHERE teacher_id=?", (id,))
        con.execute("DELETE FROM teachers WHERE id=?", (id,))

    con.commit()
    con.close()
    flash("Teacher deleted", "success")
    return redirect("/" + session["role"])


@app.route("/edit_student/<int:id>", methods=["GET", "POST"])
@login_required(["teacher", "principal", "hod"])
def edit_student(id):
    con = db()
    courses, subjects, years, semesters = get_common()
    s = con.execute("SELECT * FROM students WHERE id=?", (id,)).fetchone()

    if session.get("role") == "hod" and s["course"] != session.get("branch"):
        con.close()
        flash("You can edit only your branch students", "danger")
        return redirect("/hod")

    if request.method == "POST":
        d = request.form
        course = session.get("branch") if session.get("role") == "hod" else d["course"]
        passout = int(d["admission_year"]) + course_duration(course)

        con.execute("""UPDATE students SET name=?,email=?,roll_no=?,course=?,year=?,semester=?,admission_year=?,passout_year=?,phone=?,address=?,fees_total=?,fees_paid=? WHERE id=?""",
        (d["name"], d["email"], d["roll_no"], course, d["year"], d["semester"], int(d["admission_year"]),
         passout, d["phone"], d["address"], int(d["fees_total"]), int(d["fees_paid"]), id))
        con.commit()
        con.close()
        flash("Student updated", "success")
        return redirect("/" + session["role"])

    con.close()
    co = "".join([f"<option {'selected' if s['course']==c['name'] else ''}>{c['name']}</option>" for c in courses])
    yo = "".join([f"<option {'selected' if s['year']==y else ''}>{y}</option>" for y in years])
    so = "".join([f"<option {'selected' if s['semester']==sem else ''}>{sem}</option>" for sem in semesters])

    course_field = f"<input class='form-control' value='{session.get('branch')}' disabled>" if session.get("role") == "hod" else f"<select name='course' class='form-control'>{co}</select>"

    content = f"""
    <div class="card p-4"><h3>Edit Student</h3>
    <form method="POST" class="row g-2">
    <div class="col-md-4"><input name="name" value="{s['name']}" class="form-control"></div>
    <div class="col-md-4"><input name="email" value="{s['email']}" class="form-control"></div>
    <div class="col-md-4"><input name="roll_no" value="{s['roll_no']}" class="form-control"></div>
    <div class="col-md-3">{course_field}</div>
    <div class="col-md-3"><select name="year" class="form-control">{yo}</select></div>
    <div class="col-md-3"><select name="semester" class="form-control">{so}</select></div>
    <div class="col-md-3"><input name="admission_year" type="number" value="{s['admission_year']}" class="form-control"></div>
    <div class="col-md-3"><input name="phone" value="{s['phone']}" class="form-control"></div>
    <div class="col-md-3"><input name="fees_total" value="{s['fees_total']}" class="form-control"></div>
    <div class="col-md-3"><input name="fees_paid" value="{s['fees_paid']}" class="form-control"></div>
    <div class="col-md-12"><input name="address" value="{s['address']}" class="form-control"></div>
    <div><button class="btn btn-primary">Update</button></div>
    </form></div>
    """
    return page(content)


@app.route("/edit_teacher/<int:id>", methods=["GET", "POST"])
@login_required(["principal", "hod"])
def edit_teacher(id):
    con = db()
    courses, subjects, years, semesters = get_common()
    t = con.execute("SELECT * FROM teachers WHERE id=?", (id,)).fetchone()

    if session.get("role") == "hod" and t["course"] != session.get("branch"):
        con.close()
        flash("You can edit only your branch teachers", "danger")
        return redirect("/hod")

    if request.method == "POST":
        d = request.form
        course = session.get("branch") if session.get("role") == "hod" else d["course"]

        con.execute("""UPDATE teachers SET name=?,email=?,course=?,subject=?,phone=?,qualification=?,lectures_taken=? WHERE id=?""",
        (d["name"], d["email"], course, d["subject"], d["phone"], d["qualification"], int(d["lectures_taken"]), id))
        con.commit()
        con.close()
        flash("Teacher updated", "success")
        return redirect("/" + session["role"])

    con.close()
    co = "".join([f"<option {'selected' if t['course']==c['name'] else ''}>{c['name']}</option>" for c in courses])
    so = "".join([f"<option {'selected' if t['subject']==s['name'] else ''}>{s['name']}</option>" for s in subjects])
    course_field = f"<input class='form-control' value='{session.get('branch')}' disabled>" if session.get("role") == "hod" else f"<select name='course' class='form-control'>{co}</select>"

    content = f"""
    <div class="card p-4"><h3>Edit Teacher</h3>
    <form method="POST" class="row g-2">
    <div class="col-md-4"><input name="name" value="{t['name']}" class="form-control"></div>
    <div class="col-md-4"><input name="email" value="{t['email']}" class="form-control"></div>
    <div class="col-md-4">{course_field}</div>
    <div class="col-md-4"><select name="subject" class="form-control">{so}</select></div>
    <div class="col-md-4"><input name="phone" value="{t['phone']}" class="form-control"></div>
    <div class="col-md-4"><input name="qualification" value="{t['qualification']}" class="form-control"></div>
    <div class="col-md-4"><input name="lectures_taken" type="number" value="{t['lectures_taken']}" class="form-control"></div>
    <div><button class="btn btn-primary">Update</button></div>
    </form></div>
    """
    return page(content)


@app.route("/assign_subject", methods=["POST"])
@login_required(["teacher"])
def assign_subject():
    con = db()
    con.execute("INSERT INTO student_subjects(student_id,subject,teacher_name,created_at) VALUES(?,?,?,?)",
                (request.form["student_id"], request.form["subject"], session["name"], now()))
    con.commit()
    con.close()
    return redirect("/teacher")


@app.route("/add_marks", methods=["POST"])
@login_required(["teacher"])
def add_marks():
    d = request.form
    con = db()
    con.execute("INSERT INTO marks(student_id,subject,test_name,marks_obtained,total_marks,created_at) VALUES(?,?,?,?,?,?)",
                (d["student_id"], d["subject"], d["test_name"], int(d["marks_obtained"]), int(d["total_marks"]), now()))
    con.commit()
    con.close()
    return redirect("/teacher")


@app.route("/add_attendance", methods=["POST"])
@login_required(["teacher"])
def add_attendance():
    d = request.form
    con = db()
    con.execute("INSERT INTO attendance(student_id,subject,total_lectures,present_lectures,created_at) VALUES(?,?,?,?,?)",
                (d["student_id"], d["subject"], int(d["total_lectures"]), int(d["present_lectures"]), now()))
    con.commit()
    con.close()
    return redirect("/teacher")


@app.route("/export_students_excel")
@login_required(["principal", "teacher", "hod"])
def export_students_excel():
    from openpyxl import Workbook
    students = filtered_students()
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(["Name", "Email", "Roll No", "Course", "Year", "Semester", "Admission Year", "Passout Year", "Phone", "Address", "Fees Total", "Fees Paid", "Remaining"])
    for s in students:
        ws.append([s["name"], s["email"], s["roll_no"], s["course"], s["year"], s["semester"], s["admission_year"], s["passout_year"], s["phone"], s["address"], s["fees_total"], s["fees_paid"], s["fees_total"] - s["fees_paid"]])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="students_report.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/import_students_excel", methods=["POST"])
@login_required(["principal", "teacher", "hod"])
def import_students_excel():
    from openpyxl import load_workbook
    file = request.files["file"]
    wb = load_workbook(file)
    ws = wb.active
    con = db()
    count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            name,email,password,roll_no,course,year,semester,admission_year,phone,address,fees_total,fees_paid = row
            if session.get("role") == "hod":
                course = session.get("branch")

            if con.execute("SELECT * FROM users WHERE email=?", (email,)).fetchone():
                continue

            passout = int(admission_year) + course_duration(course)
            c = con.cursor()
            c.execute("INSERT INTO users(name,email,password,role,branch,created_at) VALUES(?,?,?,?,?,?)",
                      (name, email, generate_password_hash(str(password)), "student", course, now()))
            uid = c.lastrowid

            c.execute("""INSERT INTO students(user_id,name,email,roll_no,course,year,semester,admission_year,passout_year,phone,address,fees_total,fees_paid,created_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (uid, name, email, roll_no, course, year, semester, int(admission_year), passout, phone, address, int(fees_total), int(fees_paid), now()))
            count += 1
        except Exception:
            continue

    con.commit()
    con.close()
    flash(f"{count} students imported", "success")
    return redirect("/" + session["role"])


@app.route("/export_students_pdf")
@login_required(["principal", "teacher", "hod"])
def export_students_pdf():
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    students = filtered_students()
    output = BytesIO()
    pdf = canvas.Canvas(output, pagesize=A4)
    width, height = A4
    y = height - 50
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(40, y, "PLITMS Student Report")
    y -= 30
    pdf.setFont("Helvetica", 9)
    for s in students:
        line = f"{s['name']} | Roll: {s['roll_no']} | {s['course']} | {s['year']} | {s['semester']} | Passout: {s['passout_year']}"
        pdf.drawString(40, y, line[:120])
        y -= 18
        if y < 50:
            pdf.showPage()
            pdf.setFont("Helvetica", 9)
            y = height - 50
    pdf.save()
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="students_report.pdf", mimetype="application/pdf")


if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=5000, debug=False)


